import os, openpyxl, csv
from openpyxl.utils import get_column_letter
# CSVファイルを格納するファイルを作成
mkdir_name = os.path.basename(os.getcwd()) + '_to_csv'
print(mkdir_name)
os.makedirs(mkdir_name, exist_ok=True)

# TODO: ディレクトリのファイルをすべて読み込む
dir_list = os.listdir()
print(dir_list)
# TODO: ExcelファイルをCSVファイルに変換
for filename in dir_list:
    # TODO: CSVファイルかどうかを判定
    print('filename: {}'.format(filename))
    if not filename.endswith('.xlsx'):
        # print('￢Excelファイル')
        continue
    print(f'{filename}をCSVファイルに変換しています...')

    # TODO: Excelファイルを開く
    wb = openpyxl.load_workbook(filename, data_only=False)   # エクセルファイルからWorkbookオブジェクトを生成
    sheet_list = wb.get_sheet_names()
    sheet_num = len(sheet_list)
    # print(sheet_list)
    # print(len(sheet_list))

    # TODO: Sheetを周回して、CSVファイルを作成する
    filename_list = filename.split('.')     
    pre_filename = filename_list[0]         # '_csv.xls'以前のファイル名
    for i in range(sheet_num):
        sheet = wb.get_sheet_by_name(sheet_list[i])
        
        # TODO: 非アクティブのシートを飛ばす
        if sheet.max_column == 1 and sheet.max_row == 1 and sheet['A1'].value == None:
            continue

        # TODO: CSVファイルの作成&writerオブジェクトの作成
        csv_name = pre_filename + '_' + sheet_list[i] + '.csv'
        # print('変換後のファイル名 :' + csv_name)
        outfile = open(os.path.join(mkdir_name, csv_name), 'w', newline='')    # パスを設定し、その中で書き込みファイルを作成
        outfile_writer = csv.writer(outfile)

        # TODO: sheetを1列ずつ周回してCSVファイルへ書き込み
        cells = tuple(sheet.rows)
        # print(cells)
        for col in cells:
            col_value = []
            for j in range(len(col)):
                # print(col[j].value)
                col_value.append(col[j].value)
            # print(col_value)
            outfile_writer.writerow(col_value)    
        outfile.close()



            


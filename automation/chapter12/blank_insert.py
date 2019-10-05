#! python3
# blank_insert.py : エクセルファイルの指定行に空白行を挿入する
# CL引数：　整数(指定行), 整数(挿入行数), ファイル名

import sys, openpyxl
from openpyxl.utils import get_column_letter

# TODO: コマンドライン引数の判定
if len(sys.argv) != 4:
    print("コマンドライン引数の数が合致しません")
    exit()
if not sys.argv[1].isdigit() or not sys.argv[2].isdigit():
    print("第1引数、第2引数には数字を指定してください")
point = int(sys.argv[1])
blank = int(sys.argv[2])
filename = sys.argv[3]
print(f'{filename}の{point}行目に{blank}行の空白行を追加します。')

# TODO: スプレッドシートの内容を読み込む
wb = openpyxl.load_workbook(sys.argv[3])
sheet = wb.active
new_filename = 'blank_' + sys.argv[3]

#TODO: 書き込み用ファイルを生成し、セルオブジェクト行ごとにをリストcellsに格納する
cell_t = tuple(sheet.rows)
cells = list(cell_t)


# TODO: 最初のN行をファイルに書き込む
for i in range(point):
    row = cells[i]
    column_num = len(row)
    # row内のセルオブジェクトをVALUEへ変換する
    row_value = []
    for cell in row:
        row_value.append(cell.value)
    # print(row_value)
    for j in range(column_num):
        cell_point = get_column_letter(j+1) + str(i+1)
        # print(cell_point, row_value[j])
        sheet[cell_point] = row_value[j]

rem_row = sheet.max_row - point
# print(f'rem_row:{rem_row}')
# print(f'point:{point}')

# TODO: N+1行目を以降をM行分だけ後ろへ移動させる(行番号が大きい順で))
for i in range(point+rem_row-1, point-1, -1):
    # print(f'i:{i}')
    row = cells[i]
    column_num = len(row)
    # row内のセルオブジェクトをVALUEへ変換する
    row_value = []
    for cell in row:
        row_value.append(cell.value)
    # print(row_value)
    for j in range(column_num):
        cell_point = get_column_letter(j+1) + str(blank+i+1)
        # print(cell_point, row_value[j])
        sheet[cell_point] = row_value[j]

# TODO: 空白行の追加
for i in range(point, point + blank, 1):
    row = cells[i]
    column_num = len(row)
    for j in range(column_num):
        cell_point = get_column_letter(j+1) + str(i+1)
        # print(cell_point)
        sheet[cell_point] = None

print('完了！')
wb.save(new_filename)



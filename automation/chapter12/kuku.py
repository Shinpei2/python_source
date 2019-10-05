import openpyxl, sys
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font

# コマンドライン引数のチェック
if len(sys.argv) != 2 :
    print('CL引数の数が誤っています。')
    exit()
if sys.argv[1].isdigit() == False:
    print('第1引数に数字を指定してください。')
    exit()
num = int(sys.argv[1])

# シートの用意
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = '九九の計算'

# TODO: ウインドウ枠の作成
bold = Font(bold = True)
# 横
for j in range(num):
    # print(get_column_letter(j+2) + '1')
    sheet[get_column_letter(j+2) + '1'].value = j+1
    sheet[get_column_letter(j+2) + '1'].font = bold
# 縦
for i in range(num):
    sheet['A' + str(i+2)].value = i+1
    sheet['A' + str(i+2)].font = bold
    pass


# TODO: 計算
for i in range(num):
    column = get_column_letter(i+2)
    for j in range(num):
        # print(column + str(j+2))
        sheet[column + str(j+2)].value = (i+1) * (j+1)
        # print(str((i+1) * (j+1)))
wb.save('kuku.xlsx')
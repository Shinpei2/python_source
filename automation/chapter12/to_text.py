#! python3
# blank_insert.py : エクセルファイルの指定行に空白行を挿入する
# CL引数：　整数(指定行), 整数(挿入行数), ファイル名

import sys, openpyxl
from openpyxl.utils import get_column_letter

# TODO: コマンドライン引数の判定
if len(sys.argv) != 2:
    print("コマンドライン引数にファイル名を指定してください")
    exit()

filename = sys.argv[1]

# TODO: スプレッドシートの内容を読み込む
wb = openpyxl.load_workbook(filename)
sheet = wb.active

cell_t = tuple(sheet.columns)
cells = list(cell_t)
column = sheet.max_column
print(column)
for i in range(column):
    outfile = open('text' + str(i+1) + '.txt', 'w')
    print(f'{i+1}列目をテキストファイルへ書き写します。')
    print('テキストファイル名：' + 'text' + str(i+1) + '.txt' )
    row = len(cells[i])
    # print(row)

    # 1列分のセルをリストに格納する
    lines = list(cells[i])
    lines_val = []
    for j in range(row):
        lines_val.append(lines[j].value)
    # print(lines_val)
    # 1列の中の1行ずつをテキストファイルに書き込む
    for j in range(row):
        outfile.write(lines_val[j])
        # print(lines_val[j], 'j: ' + str(j))
        if j != row - 1:
            outfile.write('\n')
    outfile.close()
    
import openpyxl, sys
from openpyxl.utils import get_column_letter, column_index_from_string

# CL引数でエクセルファイルを指定
if len(sys.argv) != 2:
    print('CL引数にエクセルファイルを指定してください。')
    exit(1)

# 読み込みファイルの指定
wb = openpyxl.load_workbook(sys.argv[1])
sheet = wb.active

# 書き込みファイルの指定
out_wb = openpyxl.Workbook()
out_sheet = out_wb.active

# 行および列の最大値を取得
max_column = sheet.max_column
max_row = sheet.max_row
print(max_column, max_row)

# 行と列の位置を入れ替え
count = 1
for col in range(max_column):
    col_letter = get_column_letter(col+1) 
    for row in range(max_row):
        # 元のセルの位置の値を取得
        print('元のセル位置：' + col_letter + str(row+1))
        ori_value = sheet[col_letter + str(row+1)].value
        # 入替後の位置に値を格納
        re_row = column_index_from_string(col_letter)
        re_col = get_column_letter(row+1)
        print('入替後のセル位置：' + re_col + str(re_row))
        out_sheet[re_col + str(re_row)].value = ori_value


out_wb.save('rp_' + sys.argv[1])
#! python3

# to_sheet.py : カレントディレクトリ内の全てのテキストファイルを読み込み、エクセルシートに書き出す

import os, openpyxl
from openpyxl.utils import get_column_letter

# CD内の全ファイルをリストに格納する
dir_list = os.listdir()
print(dir_list)
new_filename = os.path.basename(os.getcwd()) + '_(text_to_sheet).xlsx'

# sheetオブジェクトの作成
wb = openpyxl.Workbook()
sheet = wb.active

column = 1
for filename in dir_list:
    if not filename.endswith('.txt'):
        print(f'{filename}：飛ばします')
        continue # CSVファイルでないので飛ばす
    print("シートに移しています。 " + filename + '...')
    infile = open(filename)
    infile_lines = infile.readlines()
    column_alpha = get_column_letter(column)
    sheet[column_alpha + '1'] = filename
    for i in range(len(infile_lines)):
        pre_line = infile_lines[i]
        line_split = pre_line.split('\n')
        line = line_split[0]
        cell_point = column_alpha + str(i+2)
        # print(cell_point, line)
        sheet[cell_point] = line
    column += 1

# print(new_filename)
wb.save(new_filename)

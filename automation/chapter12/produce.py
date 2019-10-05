import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet =wb.get_sheet_by_name('Sheet')

price_updates = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}

# 全ての行を走査する
for row_num in range(2, sheet.max_row +1):
    produce = sheet['A' + str(row_num)].value
    # print(produce)
    if produce in price_updates:
        sheet.cell(row=row_num, column=2).value = price_updates[produce]

wb.save('updatedProduceSales.xlsx')
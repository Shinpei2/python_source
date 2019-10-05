import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
wb = openpyxl.load_workbook('example.xlsx')    # wookbook型オブジェクト
print(type(wb))

sheets = wb.get_sheet_names()
# print(sheets)

s1 = wb.get_sheet_by_name('Sheet1')
# for i in range(1, 8,1):
#     print(str(i) + ' ' + s1.cell(row=i, column=2).value)

# print(get_column_letter(s1.max_column))
# print(column_index_from_string('C'))

# タプルの各要素：rowがまとまってる　※数字の方
cells = tuple(s1['A1':'C3'])
print(cells)

# A1 ~ C3 までの各セルの値を表示します。
for cells_obj in s1['A1':'C3']:     # tupleで中身確認してね
    for cell in cells_obj:
        print(cell.coordinate, cell.value)
    print("----END OF RUN----")